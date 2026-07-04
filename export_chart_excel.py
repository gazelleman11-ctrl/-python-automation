"""
10月販売構成.xlsx を読み込み、グラフ付きのExcelレポートを生成する。
出力: sales_report_oct.xlsx
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

EXCEL_FILE = "10月販売構成.xlsx"
OUTPUT     = "sales_report_oct.xlsx"

# ── 色定数 ──────────────────────────────────────────────
NAVY       = "0D1B2E"
BLUE_MID   = "2A78D6"
BLUE_LIGHT = "EDF1F8"
WHITE      = "FFFFFF"
GRID_LINE  = "DDE4EF"
AMBER      = "D98000"


def thin_border(sides="all"):
    s = Side(style="thin", color=GRID_LINE)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    b = Border()
    for side in sides:
        setattr(b, side, s)
    return b


def load_data() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_FILE, engine="openpyxl")
    df.columns = [c.strip() for c in df.columns]
    col_name = df.columns[0]
    df = df[~df[col_name].astype(str).str.contains("計", na=False)].reset_index(drop=True)
    df.columns = ["商品名", "販売合計数", "販売合計金額"]
    return df.sort_values("販売合計金額", ascending=False).reset_index(drop=True)


def write_excel(df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "10月販売構成"

    total_count  = df["販売合計数"].sum()
    total_amount = df["販売合計金額"].sum()

    # ── 列幅 ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    # ── タイトル行 ───────────────────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "10月 商品別販売構成レポート"
    title_cell.font      = Font(name="Meiryo UI", bold=True, size=14, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # ── KPI 帯 ───────────────────────────────────────────
    ws.merge_cells("A2:B2")
    ws.merge_cells("C2:D2")
    ws["A2"].value = f"合計売上：¥{total_amount:,}"
    ws["C2"].value = f"総販売数：{total_count:,} 個"
    ws["E2"].value = f"{len(df)} 品目"
    for cell in [ws["A2"], ws["C2"], ws["E2"]]:
        cell.font      = Font(name="Meiryo UI", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor="184F95")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    # ── テーブルヘッダー ─────────────────────────────────
    headers = ["商品名", "販売合計数", "販売合計金額", "数量比", "金額比"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font      = Font(name="Meiryo UI", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=BLUE_MID)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    ws.row_dimensions[3].height = 20

    # ── データ行 ─────────────────────────────────────────
    for i, row in df.iterrows():
        r = i + 4
        bg = WHITE if i % 2 == 0 else BLUE_LIGHT
        fill = PatternFill("solid", fgColor=bg)

        pct_cnt = row["販売合計数"]  / total_count  * 100
        pct_amt = row["販売合計金額"] / total_amount * 100

        values = [
            row["商品名"],
            row["販売合計数"],
            row["販売合計金額"],
            pct_cnt / 100,
            pct_amt / 100,
        ]
        aligns = ["left", "right", "right", "right", "right"]
        fmts   = [None, "#,##0", "¥#,##0", "0.0%", "0.0%"]

        for col, (val, align, fmt) in enumerate(zip(values, aligns, fmts), 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font      = Font(name="Meiryo UI", size=10)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal=align, vertical="center", indent=1)
            cell.border    = thin_border()
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[r].height = 18

    # ── 合計行 ───────────────────────────────────────────
    total_row = len(df) + 4
    totals = ["合計", total_count, total_amount, 1.0, 1.0]
    aligns = ["left", "right", "right", "right", "right"]
    fmts   = [None, "#,##0", "¥#,##0", "0.0%", "0.0%"]
    for col, (val, align, fmt) in enumerate(zip(totals, aligns, fmts), 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font      = Font(name="Meiryo UI", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal=align, vertical="center", indent=1)
        cell.border    = thin_border()
        if fmt:
            cell.number_format = fmt
    ws.row_dimensions[total_row].height = 20

    # ── 棒グラフ ─────────────────────────────────────────
    chart = BarChart()
    chart.type    = "bar"          # 横棒
    chart.grouping = "clustered"
    chart.title   = "10月 商品別販売金額"
    chart.y_axis.title = None
    chart.x_axis.title = "販売合計金額（円）"
    chart.legend  = None
    chart.width   = 18
    chart.height  = 12

    data_ref = Reference(ws,
                         min_col=3, max_col=3,
                         min_row=3, max_row=3 + len(df))   # row3=ヘッダー、row4〜 がデータ
    cats_ref = Reference(ws,
                         min_col=1,
                         min_row=4, max_row=3 + len(df))   # 商品名（ヘッダー除く）

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # 系列の色を濃いブルーに
    series = chart.series[0]
    series.graphicalProperties.solidFill = BLUE_MID
    series.graphicalProperties.line.solidFill = BLUE_MID

    ws.add_chart(chart, "G2")

    wb.save(OUTPUT)
    print(f"保存しました: {OUTPUT}")


def main() -> None:
    df = load_data()
    write_excel(df)


if __name__ == "__main__":
    main()
